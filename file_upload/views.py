import openpyxl
from django.shortcuts import render

# Create your views here.
def upload_file(request):
    if request.method == 'GET':
        return render(request, 'file_upload/upload_file.html')
    else:
        print("Uploading file...")
        excel_file = request.FILES["myFile"]

        if excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls'):
            with open('file/static/upload/' + excel_file.name, 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)

                wb = openpyxl.load_workbook(destination)
                
                worksheet = wb.active

                worksheet_rows = []
                for i in range(1, worksheet.max_row):
                    row_data = []
                    for col in worksheet.iter_cols(1, 11):                        
                        row_data.append(col[i].value)
                    if any(row_data):
                        worksheet_rows.append(row_data)
                            
                context = { "worksheet_rows":worksheet_rows }
